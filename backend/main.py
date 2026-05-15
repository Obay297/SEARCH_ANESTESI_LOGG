"""
backend_main.py
FastAPI application: REST endpoints and WebSocket for real-time streaming.
The frontend (index.html) communicates with this file through HTTP requests and WebSocket
"""

from __future__ import annotations

import asyncio
import io
import json as _json
import math as _math
import time as _time
import urllib.request as _urllib
from pathlib import Path
from datetime import date as _date, datetime as _datetime

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from backend_config import BASE_DIR, REPORTS_DIR
from backend_models import (
    MonitoringExportRequest,
    SourceSelectRequest,
    StartRecordingRequest,
    StopRecordingRequest,
)
from backend_sessions import SessionStore
from backend_vital_parser import (
    extract_first_valid_measurements,
    extract_report_timeseries,
)
from backend_vitalrecorder import VitalRecorderBridge

try:
    from vitaldb import VitalFile as _VitalFile  # type: ignore
except Exception:
    _VitalFile = None  # type: ignore


# Application setup 

app = FastAPI(
    title="Anesthesia Monitoring Backend",
    description="REST + WebSocket backend for the pig anesthesia monitoring system.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

#  VitalRecorder bridge per server process.
session_store  = SessionStore()
vr_bridge      = VitalRecorderBridge()
current_source = {"name": "simulation"}



VR_API_BASE = "http://127.0.0.1:14041"


VR_POLL_TRACKS = [
   
    "Solar8000/HR", "B1x5M/HR", "Bx50/HR", "IntelliVue/HR", "SNUADC/HR", "Demo/HR",
    "Solar8000/ART_SBP",  "Solar8000/ART_DBP",
    "B1x5M/ART1_SBP",     "B1x5M/ART1_DBP",
    "Bx50/ART1_SBP",      "Bx50/ART1_DBP",
    "Demo/ART_SBP",        "Demo/ART_DBP",
    "Solar8000/NIBP_SBP", "Solar8000/NIBP_DBP",
    "Solar8000/NBP_SBP",  "Solar8000/NBP_DBP",
    "Demo/NBP_SBP",        "Demo/NBP_DBP",
    "B1x5M/NBP_SBP",       "B1x5M/NBP_DBP",
    "Solar8000/BT", "Demo/BT",
    "Solar8000/PLETH_SPO2", "B1x5M/PLETH_SPO2", "Bx50/PLETH_SPO2", "Demo/PLETH_SPO2",
    "Primus/FIO2", "Solar8000/FIO2",
    "Primus/FEO2", "Solar8000/FEO2",
    "Orchestra/PPF20_RATE",   "Orchestra/PPF20_CE",
    "Orchestra/RFTN20_RATE",  "Orchestra/RFTN20_CE",
    "Orchestra/ROC_RATE",
    "Orchestra/PROPOFOL_VOL", "Orchestra/KETAMINE_VOL", "Orchestra/FENTANYL_VOL",
    "Primus/INSP_SEVO",   "Primus/EXP_SEVO",
    "Primus/INSP_DES",    "Primus/EXP_DES",
    "Solar8000/GAS2_EXPIRED", "Demo/GAS1_EXPIRED",
    "Primus/FLOW_AIR",    "Demo/FLOW_RATE",
    "Primus/PIP_MBAR",    "Solar8000/VENT_PIP",  "Demo/PIP",
    "Solar8000/VENT_TV",  "Primus/TV",            "Demo/TV",
    "Solar8000/VENT_RR",  "Solar8000/RR",         "Primus/RR_CO2",  "Demo/RR", "Demo/RR_CO2",
    "Solar8000/VENT_MV",  "Primus/MV",
    "Primus/PEEP_MBAR",   "Demo/PEEP",
    "Solar8000/ETCO2", "Primus/ETCO2", "Demo/ETCO2",
    "BIS/BIS",
]




def _resolve_latest_vital_file() -> str | None:
   
    vital_file = vr_bridge.last_found_vital_file
    if not vital_file:
        newest     = vr_bridge.find_newest_vital_file()
        vital_file = str(newest) if newest else None
    return vital_file


def _zero_measurements() -> dict
    return {
        "pulse":        0,
        "tbp":          "0/0",
        "temp":         0,
        "o2_primary":   0,
        "propofol":     0,
        "ketamin":      0,
        "fentanyl":     0,
        "isofluran":    0,
        "flow":         0,
        "o2_secondary": 0,
        "pmax":         0,
        "vt":           0,
        "frequency":    0,
        "mv":           0,
        "peep":         0,
        "etco2":        0,
        "bis":          0,
    }


def _fetch_vr_values_sync() -> dict | None:
   
    try:
        trks = ",".join(_VR_POLL_TRACKS)
        url  = f"{VR_API_BASE}/api/vals?trks={trks}"
        req  = _urllib.Request(url, headers={"Accept": "application/json"})
        with _urllib.urlopen(req, timeout=1) as resp:
            data = _json.loads(resp.read())

        
        if isinstance(data, dict):
            return data
        if isinstance(data, list):
            return {
                item["trk"]: item["val"]
                for item in data
                if "trk" in item and "val" in item
            }
    except Exception:
        pass
    return None


async def _try_read_vitalrecorder_http() -> dict | None:
    
    return await asyncio.to_thread(_fetch_vr_values_sync)


def _fetch_from_growing_vital_file_sync() -> dict | None:
  f {track_name: latest_float_value}, or None on failure.
   
    if _VitalFile is None:
        return None
    try:
        newest = vr_bridge.find_newest_vital_file()
        if not newest or not newest.exists():
            return None

        # Ignore files not modified in the past 30 seconds.
        if _time.time() - newest.stat().st_mtime > 30:
            return None

        vf     = _VitalFile(str(newest))
        result: dict = {}

        for trk_name, trk in (getattr(vf, "trks", None) or {}).items():
            for rec in reversed(getattr(trk, "recs", None) or []):
                vals = rec.get("val") if isinstance(rec, dict) else getattr(rec, "val", None)
                if vals is None:
                    continue
                it = reversed(list(vals)) if hasattr(vals, "__iter__") else [vals]
                for v in it:
                    try:
                        fv = float(v)
                        if _math.isfinite(fv):
                            result[trk_name] = fv
                            break
                    except (TypeError, ValueError):
                        pass
                if trk_name in result:
                    break

        return result or None
    except Exception:
        return None
async def _try_read_growing_vital_file() -> dict | None:
    return await asyncio.to_thread(_fetch_from_growing_vital_file_sync)

# A method for all vital sign
def _map_vr_tracks_to_measurements(raw: dict) -> dict:
    def _get(*keys):
        for key in keys:
            value = raw.get(key)
            if value is not None:
                try:
                    fv = float(value)
                    if _math.isfinite(fv):
                        return round(fv, 1)
                except (TypeError, ValueError):
                    pass
        return 0

    sbp = _get(
        "Solar8000/ART_SBP",  "B1x5M/ART1_SBP",  "Bx50/ART1_SBP",  "Demo/ART_SBP",
        "Solar8000/NIBP_SBP", "Solar8000/NBP_SBP", "Demo/NBP_SBP",   "B1x5M/NBP_SBP",
    )
    dbp = _get(
        "Solar8000/ART_DBP",  "B1x5M/ART1_DBP",  "Bx50/ART1_DBP",  "Demo/ART_DBP",
        "Solar8000/NIBP_DBP", "Solar8000/NBP_DBP", "Demo/NBP_DBP",   "B1x5M/NBP_DBP",
    )

    return {
        "pulse":        _get("Solar8000/HR",          "B1x5M/HR",          "Bx50/HR",         "Demo/HR",         "IntelliVue/HR", "SNUADC/HR"),
        "tbp":          f"{sbp}/{dbp}" if sbp or dbp else "0/0",
        "temp":         _get("Solar8000/BT",           "Demo/BT"),
        "o2_primary":   _get("Solar8000/PLETH_SPO2",   "B1x5M/PLETH_SPO2",  "Bx50/PLETH_SPO2", "Demo/PLETH_SPO2"),
        "o2_secondary": _get("Primus/FEO2",            "Solar8000/FEO2"),
        "propofol":     _get("Orchestra/PPF20_RATE",   "Orchestra/PPF20_CE", "Orchestra/PROPOFOL_VOL"),
        "ketamin":      _get("Orchestra/ROC_RATE",     "Orchestra/KETAMINE_VOL"),
        "fentanyl":     _get("Orchestra/RFTN20_RATE",  "Orchestra/RFTN20_CE","Orchestra/FENTANYL_VOL"),
        "isofluran":    _get("Primus/INSP_SEVO",       "Primus/EXP_SEVO",    "Primus/INSP_DES", "Primus/EXP_DES",  "Solar8000/GAS2_EXPIRED", "Demo/GAS1_EXPIRED"),
        "flow":         _get("Primus/FLOW_AIR",        "Demo/FLOW_RATE"),
        "pmax":         _get("Primus/PIP_MBAR",        "Solar8000/VENT_PIP", "Demo/PIP"),
        "vt":           _get("Solar8000/VENT_TV",      "Primus/TV",          "Demo/TV"),
        "frequency":    _get("Solar8000/VENT_RR",      "Solar8000/RR",       "Primus/RR_CO2",   "Demo/RR",         "Demo/RR_CO2"),
        "mv":           _get("Solar8000/VENT_MV",      "Primus/MV"),
        "peep":         _get("Primus/PEEP_MBAR",       "Demo/PEEP"),
        "etco2":        _get("Solar8000/ETCO2",        "Primus/ETCO2",       "Demo/ETCO2"),
        "bis":          _get("BIS/BIS"),
    }


# REST endpoints

@app.get("/health")
async def health():
    """Simple health-check endpoint used by the frontend to verify the backend is running."""
    return {"ok": True}


@app.post("/source/select") # Switches active data source (simulation / live / file)
async def select_source(payload: SourceSelectRequest):
    current_source["name"] = payload.source
    return {"ok": True, "source": payload.source}


@app.post("/recording/start")
async def start_recording(payload: StartRecordingRequest):
  
    patient_payload = payload.patient.model_dump() if payload.patient else {}
    session         = session_store.create_session(patient_payload)

    patient_id   = (patient_payload.get("id") or "").strip()
    patient_date = (patient_payload.get("date") or "").strip()

    if patient_id:
        date_part    = patient_date[:10].replace("-", "") if patient_date else _date.today().strftime("%Y%m%d")
        desired_stem = f"{patient_id}_{date_part}"
    else:
        desired_stem = _datetime.now().strftime("recording_%Y%m%d_%H%M")

    vr_result = vr_bridge.start(desired_stem=desired_stem)

    return {
        "ok":        True,
        "source":    current_source["name"],
        "sessionId": session["session_id"],
        "bridge":    vr_result,
    }


@app.post("/recording/stop")
async def stop_recording(payload: StopRecordingRequest):
    session = session_store.get_current()
    if not session:
        return {"ok": False, "error": "No active session"}

    events_data = []
    for item in payload.events:
        if hasattr(item, "model_dump"):
            events_data.append(item.model_dump())
        else:
            events_data.append(item.dict())

    session_store.save_events(events_data)
    vr_result = vr_bridge.stop()

    vital_file         = vr_result.get("vital_file")
    first_valid_report = None
    timeline_report    = None

    if vital_file:
        try:
            first_valid_report = extract_first_valid_measurements(vital_file)
            timeline_report    = extract_report_timeseries(vital_file)
        except Exception as exc:
            return {
                "ok":        False,
                "sessionId": session["session_id"],
                "bridge":    vr_result,
                "error":     str(exc),
            }

    return {
        "ok":               True,
        "sessionId":        session["session_id"],
        "bridge":           vr_result,
        "reportData":       timeline_report,
        "firstValidReport": first_valid_report,
    }


@app.get("/analysis/latest")
async def analysis_latest():
    
    vital_file = _resolve_latest_vital_file()
    if not vital_file:
        return {"ok": False, "error": "No .vital file found"}

    try:
        report = extract_first_valid_measurements(vital_file)
    except Exception as exc:
        return {
            "ok":   False,
            "file": Path(vital_file).name,
            "error": str(exc),
        }

    return {
        "ok":     True,
        "file":   Path(vital_file).name,
        "path":   vital_file,
        "report": report,
    }


@app.get("/reports/{filename}")
async def get_report(filename: str):
     return FileResponse(REPORTS_DIR / filename, media_type="text/html")


@app.get("/record/latest") # Loads latest record + returns snapshot + timeline report for UI
async def load_latest_record():
   
    vital_file = _resolve_latest_vital_file()
    if not vital_file:
        return {"ok": False, "error": "No .vital file found"}

    try:
        first_valid_report = extract_first_valid_measurements(vital_file)
        timeline_report    = extract_report_timeseries(vital_file)
    except Exception as exc:
        return {
            "ok":    False,
            "file":  Path(vital_file).name,
            "error": str(exc),
        }

    vr_bridge.last_found_vital_file = vital_file

    # Prefer last real measurement (has all track values); fall back to snapshot.
    snapshot: dict = {}
    last_data = (timeline_report or {}).get("lastMeasurement", {}).get("data", {})
    if any(isinstance(v, (int, float)) and v != 0 for v in last_data.values()):
        snapshot = last_data
    else:
        snapshot = first_valid_report.get("snapshot", {})

    return {
        "ok":             True,
        "file":           Path(vital_file).name,
        "path":           vital_file,
        "snapshot":       snapshot,
        "reportData":     timeline_report,
        "timelineReport": timeline_report,
    }


@app.get("/debug/vital-file")  #available tracks and mapping status per signal
async def debug_vital_file():
  
    try:
        from vitaldb import VitalFile as _VF  # noqa: PLC0415
        from backend_vital_parser import (  # noqa: PLC0415
            _get_available_tracks,
            _find_track_name,
            LEGACY_TRACK_CANDIDATES,
            _read_track_direct,
        )

        vital_file = _resolve_latest_vital_file()
        if not vital_file:
            return {"ok": False, "error": "No .vital file found"}

        vf     = _VF(vital_file)
        tracks = _get_available_tracks(vf)
        found: dict = {}

        for key, candidates in LEGACY_TRACK_CANDIDATES.items():
            matched = _find_track_name(vf, candidates)
            if matched:
                arr  = _read_track_direct(vf, matched)
                last = float(arr[-1]) if arr.size > 0 else None
                found[key] = {"track": matched, "samples": int(arr.size), "last": last}
            else:
                found[key] = {"track": None, "candidates_tried": candidates}

        return {"ok": True, "file": Path(vital_file).name, "all_tracks": tracks, "found": found}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@app.get("/debug/vr-tracks") #compares HTTP API vs file-based data sources
async def debug_vr_tracks():
   
    raw_http = await _try_read_vitalrecorder_http()
    raw_file = await _try_read_growing_vital_file()
    return {
        "http_api_tracks":   sorted(raw_http.keys()) if raw_http else [],
        "vital_file_tracks": sorted(raw_file.keys()) if raw_file else [],
        "mapped":            _map_vr_tracks_to_measurements(raw_http or raw_file or {}),
    }


@app.post("/export/monitoring-xlsx")  # patient info, monitoring rows, and events
async def export_monitoring_xlsx(payload: MonitoringExportRequest):
   
    try:
        import openpyxl 
    except ImportError:
        from fastapi import HTTPException  
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    wb = openpyxl.Workbook()


    ws_meta       = wb.active
    ws_meta.title = "Session Info"
    patient       = payload.patient or {}
    ws_meta.append(["Field", "Value"])
    for key, value in (patient.items() if hasattr(patient, "items") else []):
        ws_meta.append([key, value])
    ws_meta.append([])
    ws_meta.append(["Source", payload.source or ""])

        ws_data       = wb.create_sheet("Monitoring Data")
        if payload.rows:
        headers = list(payload.rows[0].keys())
        ws_data.append(headers)
        for row in payload.rows:
            ws_data.append([row.get(h, "") for h in headers])

 
    ws_events = wb.create_sheet("Events")
    ws_events.append(["Time", "Event"])
    for event in (payload.events or []):
        ws_events.append([event.get("time", ""), event.get("text", "")])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=monitoring_export.xlsx"},
    )




@app.websocket("/live") # WebSocket endpoint for real-time streaming of vital measurements 
async def live_socket(websocket: WebSocket):
  
    await websocket.accept()

    try:
        while True:
            raw = await _try_read_vitalrecorder_http()

            if not raw:
                raw = await _try_read_growing_vital_file()

            measurements = (
                _map_vr_tracks_to_measurements(raw)
                if raw
                else _zero_measurements()
            )

            await websocket.send_json({"measurements": measurements})
            await asyncio.sleep(1)

    except (WebSocketDisconnect, Exception):
        pass

app.mount("/", StaticFiles(directory=str(BASE_DIR), html=True), name="static")  # Serves frontend static files (HTML/CSS/JS)


if __name__ == "__main__":
    import uvicorn # runs uvicorn server
    uvicorn.run("backend_main:app", host="127.0.0.1", port=8001, reload=False)
